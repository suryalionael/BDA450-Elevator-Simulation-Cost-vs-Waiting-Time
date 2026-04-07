"""
data_input.py
=============
Reads OnCounts.xlsx and OffCounts.xlsx directly and generates per-passenger
arrival events using a non-homogeneous Poisson process (piecewise-constant
rates).

Input files
-----------
OnCounts.xlsx  – passengers *boarding* per 15-min block per floor
OffCounts.xlsx – passengers *alighting* (used only for reference; destinations
                 are assigned probabilistically via DEST_PROBS below)

Both files share the same layout:
  Row 0  : header  (None, 'time', 0, '1', '2', '3', '4')
  Rows 1+: one row per 15-min block; 'time' column uses interval notation
            e.g. '(8.25, 8.5]' in decimal hours.

Arrival generation
------------------
Within each 15-minute (900 s) block:
  * Expected count = num_on * traffic_scale.
  * Actual count ~ Poisson(expected).
  * Arrival times drawn uniformly within the block (piecewise-constant
    Poisson process, standard approach).
  * Destination sampled from DEST_PROBS[origin_floor].
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from typing import List, Dict

# ─────────────────────────────────────────────
# Paths to the source Excel files
# ─────────────────────────────────────────────
ON_COUNTS_PATH  = "OnCounts.xlsx"
OFF_COUNTS_PATH = "OffCounts.xlsx"

# ─────────────────────────────────────────────
# Configurable destination-floor probabilities
# P(dest | origin) – must sum to 1 for each origin
# ─────────────────────────────────────────────
DEST_PROBS: Dict[int, Dict[int, float]] = {
    # From parking (floor 0) people mostly go to business floors
    0: {0: 0.00, 1: 0.30, 2: 0.25, 3: 0.25, 4: 0.20},
    # From floor 1 people go up or sometimes down to parking
    1: {0: 0.20, 1: 0.00, 2: 0.30, 3: 0.30, 4: 0.20},
    # From floor 2
    2: {0: 0.20, 1: 0.20, 2: 0.00, 3: 0.30, 4: 0.30},
    # From floor 3
    3: {0: 0.25, 1: 0.25, 2: 0.25, 3: 0.00, 4: 0.25},
    # From floor 4
    4: {0: 0.25, 1: 0.25, 2: 0.25, 3: 0.25, 4: 0.00},
}

# Pre-convert to parallel lists for fast numpy sampling
_DEST_ARRAYS: Dict[int, tuple] = {
    floor: (list(probs.keys()), list(probs.values()))
    for floor, probs in DEST_PROBS.items()
}


# ─────────────────────────────────────────────
# Excel reader
# ─────────────────────────────────────────────

def _read_excel_counts(path: str) -> pd.DataFrame:
    """
    Parse one of the count Excel files into a long-format DataFrame with
    columns [time_block_start (int, seconds), floor (int), count (int)].

    The 'time' column holds interval strings like '(8.25, 8.5]' in decimal
    hours; we extract the lower bound and convert to seconds.
    """
    wb   = load_workbook(path, read_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]                        # (None, 'time', 0, '1', '2', '3', '4')
    floors = [int(h) for h in header[2:]]  # [0, 1, 2, 3, 4]

    records = []
    for row in rows[1:]:
        time_label = row[1]                 # e.g. '(8.25, 8.5]'
        lo_h = float(time_label.split(",")[0].strip("(").strip())
        block_start = int(lo_h * 3600)     # decimal hours → seconds

        for floor, cnt in zip(floors, row[2:]):
            records.append({
                "time_block_start": block_start,
                "floor":            floor,
                "count":            int(cnt) if cnt else 0,
            })

    return pd.DataFrame(records)


# ─────────────────────────────────────────────
# Passenger helper
# ─────────────────────────────────────────────

def _make_passenger(pid: int, origin: int, dest: int, arrival: float) -> dict:
    return {
        "id":               pid,
        "origin_floor":     origin,
        "destination_floor": dest,
        "arrival_time":     arrival,
        "board_time":       None,
        "arrival_dest_time": None,
    }


# ─────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────

def load_passengers(
    on_path:       str   = ON_COUNTS_PATH,
    off_path:      str   = OFF_COUNTS_PATH,
    traffic_scale: float = 1.0,
    rng: np.random.Generator = None,
) -> List[dict]:
    """
    Read OnCounts.xlsx (and optionally OffCounts.xlsx) and return a list of
    passenger dicts sorted by arrival_time (seconds since midnight).

    Parameters
    ----------
    on_path       : path to OnCounts.xlsx
    off_path      : path to OffCounts.xlsx (currently unused in generation
                    but kept for signature symmetry / future use)
    traffic_scale : multiplier on expected counts  (0.8 / 1.0 / 1.2)
    rng           : numpy random Generator; created internally if None

    Returns
    -------
    List of passenger dicts with timing fields initialised to None.
    Passengers with origin == destination are dropped (no elevator needed).
    """
    if rng is None:
        rng = np.random.default_rng()

    on_df = _read_excel_counts(on_path)

    passengers: List[dict] = []
    pid = 0
    block_duration = 900  # seconds per 15-min block

    for _, row in on_df.iterrows():
        t_start  = int(row["time_block_start"])
        floor    = int(row["floor"])
        expected = float(row["count"]) * traffic_scale

        # Actual count ~ Poisson(expected)
        count = rng.poisson(expected)
        if count == 0:
            continue

        # Uniform arrivals within the block
        arrival_times = rng.uniform(t_start, t_start + block_duration, size=count)

        dest_floors, dest_probs = _DEST_ARRAYS[floor]
        destinations = rng.choice(dest_floors, size=count, p=dest_probs)

        for arr, dest in zip(arrival_times, destinations):
            dest = int(dest)
            if dest == floor:
                continue  # same-floor trip → no elevator
            passengers.append(_make_passenger(pid, floor, dest, float(arr)))
            pid += 1

    passengers.sort(key=lambda p: p["arrival_time"])
    return passengers
