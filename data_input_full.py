"""
data_input.py
=============
Reads OnCounts.xlsx and OffCounts.xlsx and generates per-passenger
arrival events using a non-homogeneous Poisson process (piecewise-constant
rates).

Input files
-----------
OnCounts.xlsx  – passengers *boarding* per 15-min block per floor
OffCounts.xlsx – passengers *alighting* per 15-min block per floor

Both files share the same layout:
  Row 0  : header  (None, 'time', 0, '1', '2', '3', '4')
  Rows 1+: one row per 15-min block; 'time' column uses interval notation
            e.g. '(8.25, 8.5]' in decimal hours.

Destination assignment (data-driven, time-varying)
--------------------------------------------------
For each 15-minute block t and origin floor i, destination probabilities
are derived directly from that block's OffCounts:

    P(dest = j | origin = i, block = t)  ∝  OffCounts[j, t],   j ≠ i

This means destinations reflect actual measured alighting patterns and
vary across the day — e.g. evening blocks naturally produce more downward
trips because parking and ground-floor off-counts are higher then.

Fallback: if a block has zero off-counts across all candidate destination
floors (e.g. very early morning), the all-day aggregate OffCounts are used
instead.  This guarantees every passenger is always assigned a valid
destination without any hardcoded assumptions.

Arrival generation
------------------
Within each 15-minute (900 s) block:
  * Expected count = num_on * traffic_scale.
  * Actual count ~ Poisson(expected).
  * Arrival times drawn uniformly within the block (standard
    piecewise-constant Poisson process).
  * Passengers with origin == destination are dropped (no elevator needed).
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from typing import List, Dict, Tuple, Optional

ON_COUNTS_PATH  = "OnCounts.xlsx"
OFF_COUNTS_PATH = "OffCounts.xlsx"

BLOCK_DURATION: int = 900   # seconds per 15-min block


# ─────────────────────────────────────────────
# Excel reader  (shared for both files)
# ─────────────────────────────────────────────

def _read_excel_counts(path: str) -> pd.DataFrame:
    """
    Parse one count Excel file into a tidy long-format DataFrame:
        columns: time_block_start (int, seconds), floor (int), count (int)

    The 'time' column contains interval strings like '(8.25, 8.5]' in
    decimal hours; we extract the lower bound and convert to seconds.
    """
    wb   = load_workbook(path, read_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]                        # (None, 'time', 0, '1', '2', ...)
    floors = [int(h) for h in header[2:]]  # [0, 1, 2, 3, 4]

    records = []
    for row in rows[1:]:
        time_label  = row[1]               # e.g. '(8.25, 8.5]'
        lo_h        = float(time_label.split(",")[0].strip("( "))
        block_start = int(lo_h * 3600)    # decimal hours → seconds

        for floor, cnt in zip(floors, row[2:]):
            records.append({
                "time_block_start": block_start,
                "floor":            floor,
                "count":            int(cnt) if cnt else 0,
            })

    return pd.DataFrame(records)


# ─────────────────────────────────────────────
# Destination-probability builder
# ─────────────────────────────────────────────

# Type alias: maps origin_floor → (dest_floor_list, probability_list)
_DestMap = Dict[int, Tuple[List[int], List[float]]]


def _build_dest_probs(
    off_df: pd.DataFrame,
    all_floors: List[int],
) -> Tuple[Dict[int, _DestMap], _DestMap]:
    """
    Derive destination probabilities from OffCounts data.

    For each time block t and origin floor i:
        P(dest = j | origin = i, block = t)  ∝  OffCounts[j, t],   j ≠ i

    When a block has zero alighting counts across all candidate
    destinations the all-day aggregate is used as a fallback.

    Returns
    -------
    block_probs  : {time_block_start -> {origin_floor -> (floors, probs)}}
    global_probs : {origin_floor -> (floors, probs)}  — all-day fallback
    """

    def _normalise(origin: int, off_counts: pd.Series,
                   fallback_to: Optional[pd.Series] = None
                   ) -> Tuple[List[int], List[float]]:
        """
        Build (dest_floors, probs) for one origin from a Series of off counts.
        Falls back to fallback_to if all counts are zero.
        """
        dests   = [f for f in all_floors if f != origin]
        weights = [float(off_counts.get(d, 0.0)) for d in dests]

        if sum(weights) == 0:
            if fallback_to is not None:
                weights = [float(fallback_to.get(d, 1.0)) for d in dests]
            else:
                weights = [1.0] * len(dests)   # pure uniform — last resort

        total = sum(weights)
        return dests, [w / total for w in weights]

    # ── All-day aggregate (used as per-block fallback) ─────────────────
    global_off = off_df.groupby("floor")["count"].sum()
    global_probs: _DestMap = {
        origin: _normalise(origin, global_off)
        for origin in all_floors
    }

    # ── Per-block probs ────────────────────────────────────────────────
    block_probs: Dict[int, _DestMap] = {}
    for t, grp in off_df.groupby("time_block_start"):
        off_counts = grp.set_index("floor")["count"]
        block_probs[int(t)] = {
            origin: _normalise(origin, off_counts, fallback_to=global_off)
            for origin in all_floors
        }

    return block_probs, global_probs


# ─────────────────────────────────────────────
# Passenger factory
# ─────────────────────────────────────────────

def _make_passenger(pid: int, origin: int, dest: int, arrival: float) -> dict:
    return {
        "id":                pid,
        "origin_floor":      origin,
        "destination_floor": dest,
        "arrival_time":      arrival,
        "board_time":        None,
        "arrival_dest_time": None,
        "pass_count":        0,
    }


# ─────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────

def load_passengers(
    on_path:       str                    = ON_COUNTS_PATH,
    off_path:      str                    = OFF_COUNTS_PATH,
    traffic_scale: float                  = 1.0,
    rng:           np.random.Generator   = None,
) -> List[dict]:
    """
    Read OnCounts.xlsx and OffCounts.xlsx and return a list of passenger
    dicts sorted by arrival_time (seconds since midnight).

    Parameters
    ----------
    on_path       : path to OnCounts.xlsx
    off_path      : path to OffCounts.xlsx
    traffic_scale : multiplier on expected arrival counts (0.8 / 1.0 / 1.2)
    rng           : numpy random Generator; one is created if not supplied

    Returns
    -------
    List of passenger dicts, sorted by arrival_time.
    Trips with origin == destination are excluded (no elevator needed).
    """
    if rng is None:
        rng = np.random.default_rng()

    on_df  = _read_excel_counts(on_path)
    off_df = _read_excel_counts(off_path)

    all_floors = sorted(set(on_df["floor"].unique()) | set(off_df["floor"].unique()))

    # Build data-driven destination probabilities from OffCounts
    block_probs, global_probs = _build_dest_probs(off_df, all_floors)

    passengers: List[dict] = []
    pid = 0

    for _, row in on_df.iterrows():
        t_start  = int(row["time_block_start"])
        floor    = int(row["floor"])
        expected = float(row["count"]) * traffic_scale

        # Actual arrivals ~ Poisson(expected)
        count = rng.poisson(expected)
        if count == 0:
            continue

        # Uniform arrival times within the block
        arrival_times = rng.uniform(t_start, t_start + BLOCK_DURATION, size=count)

        # Destination probs: block-specific with global fallback
        dest_floors, dest_probs = block_probs.get(t_start, global_probs).get(
            floor, global_probs[floor]
        )

        destinations = rng.choice(dest_floors, size=count, p=dest_probs)

        for arr, dest in zip(arrival_times, destinations):
            dest = int(dest)
            if dest == floor:
                continue    # safety guard (should not occur given probs exclude self)
            passengers.append(_make_passenger(pid, floor, dest, float(arr)))
            pid += 1

    passengers.sort(key=lambda p: p["arrival_time"])
    return passengers
